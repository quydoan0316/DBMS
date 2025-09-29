from openpyxl import load_workbook
import io
import base64
from pathlib import Path
import pydicom
from fastapi import FastAPI, HTTPException, Query
from bson import ObjectId
from fastapi.responses import StreamingResponse
from PIL import Image
import numpy as np
from fastapi.middleware.cors import CORSMiddleware
import re
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from zipfile import ZipFile
import tempfile
import os
import time
from pydantic import BaseModel
from fastapi.responses import FileResponse
from database import files_collection, fs_bucket, chunks_collection, patients_diagnosis, patients_medicalHistory
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from pathlib import Path


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DATA_DIR = Path(r"C:\Users\ADMIN\Documents\BK_Document\Năm-IV\251\DBMS\EW\Clone\01_MRI_Data")

VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh")

class MedicalHistoryIn(BaseModel):
    patient_id: int
    progress_note: str

def serialize_history(doc):
    d = doc.get("date")
    if isinstance(d, datetime):
        if d.tzinfo is None:
            d = d.replace(tzinfo=timezone.utc)
        d_vn = d.astimezone(VN_TZ)
        date_out = d_vn.strftime("%d/%m/%Y %H:%M:%S")
    else:
        date_out = ""
    return {
        "patient_id": doc.get("patient_id"),
        "progress_note": doc.get("progress_note"),
        "date": date_out, 
    }

@app.get("/")
async def root():
    return {"message": "Hello World"}

@app.get("/app")
async def serve_app():
    html_path = Path(__file__).parent / "frontend.html"
    if not html_path.exists():
        raise HTTPException(status_code=404, detail="frontend.html not found")
    return FileResponse(str(html_path))

@app.post("/medical-history")
async def create_medical_history(entry: MedicalHistoryIn):
    doc = {
        "patient_id": entry.patient_id,
        "progress_note": entry.progress_note,
        "date": datetime.now(timezone.utc),
    }
    try:
        await patients_medicalHistory.insert_one(doc)
        return {"message": "Inserted", "data": serialize_history(doc)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/medical-history/{patient_id}")
async def get_medical_history(patient_id: int):
    try:
        cursor = patients_medicalHistory.find({"patient_id": patient_id}).sort("date", -1)
        items = []
        async for doc in cursor:
            items.append(serialize_history(doc))
        return {"patient_id": patient_id, "count": len(items), "items": items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================================
# Hàm build cây thư mục nhiều cấp
# ==========================================================
def insert_into_tree(tree, path_parts, file_entry):
    """
    Đệ quy chèn file_entry vào cây thư mục.
    path_parts: danh sách tên folder (không bao gồm patient_id, không bao gồm tên file).
    """
    if not path_parts:
        return

    folder = path_parts[0]
    if len(path_parts) == 1:
        # Folder cuối cùng chứa file → list ảnh
        tree.setdefault(folder, []).append(file_entry)
    else:
        # Còn thư mục con → đi tiếp
        tree.setdefault(folder, {})
        insert_into_tree(tree[folder], path_parts[1:], file_entry)

# ==========================================================
# Các API cơ bản
# ==========================================================
@app.get("/")
async def root():
    return {"message": "Hello World"}


@app.get("/files/{id}")
async def get_file(id: str):
    file = await files_collection.find_one({"_id": ObjectId(id)})
    if file:
        file["_id"] = str(file["_id"]) 
        return file
    return {"error": "File not found"}


@app.get("/chunks/{id}")
async def get_chunk_raw(id: str):
    chunk = await chunks_collection.count_documents({"files_id": ObjectId(id)})
    return StreamingResponse(io.BytesIO(chunk["data"]), media_type="application/octet-stream")

# ==========================================================
# Upload dữ liệu
# ==========================================================

@app.get("/upload")
async def upload_all_dicom_files():
    uploaded = []
    excel_file = DATA_DIR / "Radiologists Report.xlsx"
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    
    start_time = time.time()

    for i in range(0, 575):
        row = sheet.cell(row=i+2, column=1).value
        diagnosis = sheet.cell(row=i+2, column=2).value    
        regimen = sheet.cell(row=i+2, column=3).value 
        await patients_diagnosis.insert_one({
            "patient_id": row,
            "diagnosis": diagnosis,
            "regimen": regimen,
            "root": str(row).zfill(4)
        })
        print(f"Uploading patient {row}")
        folder = list(DATA_DIR.rglob("*"))[i]
        for file_path in folder.rglob("*.ima"):
            try:
                ds = pydicom.dcmread(str(file_path))
                metadata = {
                    elem.keyword: str(elem.value)
                    for elem in ds if elem.VR != "SQ"
                }
                with open(file_path, "rb") as f:
                    file_id = await fs_bucket.upload_from_stream(file_path.name, f, metadata=metadata)
                await files_collection.find_one_and_update(
                    {"_id": file_id},
                    {"$set": {"filename": str(file_path.relative_to(DATA_DIR))}}
                )
                uploaded.append(str(file_path))
            except Exception as e:
                print(f"Failed {file_path}: {e}")
    
    elapsed_time = time.time() - start_time

    return {"elapsed_time_seconds": elapsed_time}

# ==========================================================
# API search theo keyword
# ==========================================================
@app.get("/search_keyword_with_images")
async def search_keyword_with_images(
    keyword: str = Query(..., description="Keyword để tìm chẩn đoán"),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1)
):
    skip = (page - 1) * page_size
    cursor = patients_diagnosis.find(
        {"diagnosis": {"$regex": keyword, "$options": "i"}}
    ).skip(skip).limit(page_size)

    patient_list = await cursor.to_list(length=page_size)
    if not patient_list:
        raise HTTPException(status_code=404, detail="No patient found for this keyword")

    results = []
    for patient in patient_list:
        processed = await process_patient_files(patient)
        results.append(processed)

    return {
        "page": page,
        "page_size": page_size,
        "patients": results
    }

# ==========================================================
# Xử lý chung cho 1 patient
# ==========================================================
async def process_patient_files(patient):
    root = patient["root"]
    patient_id = patient.get("patient_id", "")
    diagnosis = patient.get("diagnosis", "")
    regimen = patient.get("regimen", "")

    file_cursor = files_collection.find({"filename": {"$regex": f"^{root}"}})
    files = await file_cursor.to_list(length=None)
    folders = {}

    for file in files:
        try:
            file_id = file["_id"]
            stream = await fs_bucket.open_download_stream(file_id)
            dicom_bytes = await stream.read()

            ds = pydicom.dcmread(io.BytesIO(dicom_bytes))
            pixel_array = ds.pixel_array
            pixel_array = (
                (pixel_array - np.min(pixel_array)) /
                (np.max(pixel_array) - np.min(pixel_array)) * 255
            ).astype(np.uint8)

            image = Image.fromarray(pixel_array)
            img_byte_arr = io.BytesIO()
            image.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            img_base64 = base64.b64encode(img_byte_arr.read()).decode("utf-8")

            metadata = {
                "PatientID": patient_id,
                "PatientBirthDate": getattr(ds, "PatientBirthDate", ""),
                "PatientSex": getattr(ds, "PatientSex", ""),
                "PatientAge": getattr(ds, "PatientAge", ""),
                "PatientSize": getattr(ds, "PatientSize", ""),
                "PatientWeight": getattr(ds, "PatientWeight", ""),
                "PatientIdentityRemoved": getattr(ds, "PatientIdentityRemoved", ""),
                "BodyPartExamined": getattr(ds, "BodyPartExamined", "")
            }

            filepath = Path(file["filename"])
            parts = filepath.parts 
            relative_parts = parts[1:] 
            file_entry = {
                "filename": file["filename"],
                "metadata": metadata,
                "image_base64": img_base64
            }
            insert_into_tree(folders, list(relative_parts[:-1]), file_entry)

        except Exception as e:
            print(f"Error processing file {file['_id']}: {e}")
            continue

    return {
        "patient_id": patient_id,
        "diagnosis": diagnosis,
        "regimen": regimen,
        "folders": folders
    }

# ==========================================================
# API search theo patient_id 
# ==========================================================
@app.get("/search_patient_by_id")
async def search_patient_by_id(
    patient_id: int = Query(..., description="Patient ID để tìm")
):
    patient = await patients_diagnosis.find_one({"patient_id": patient_id})
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    result = await process_patient_files(patient)
    return result
# ==========================================================
# API upload bệnh nhân từ folder zip 
# ==========================================================
@app.post("/add_patient_from_folder")
async def add_patient_from_folder(
    diagnosis: str = Form(...),
    regimen: str = Form(...),
    zip_file: UploadFile = File(...)
):
    try:
        
        last_patient = await patients_diagnosis.find_one(sort=[("patient_id", -1)])
        new_id = int(last_patient["patient_id"]) + 1 if last_patient else 1
        patient_id = str(new_id).zfill(4) 

        tmp_dir = tempfile.mkdtemp()
        zip_path = os.path.join(tmp_dir, zip_file.filename)
        with open(zip_path, "wb") as f:
            f.write(await zip_file.read())

        with ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(tmp_dir)

        await patients_diagnosis.insert_one({
            "patient_id": new_id,
            "diagnosis": diagnosis,
            "regimen": regimen,
            "root": patient_id
        })

        uploaded_files = []

        for root, _, files in os.walk(tmp_dir):
            for file_name in files:
                if not file_name.lower().endswith(".ima"):
                    continue

                file_path = os.path.join(root, file_name)

                try:
                    ds = pydicom.dcmread(file_path)
                    metadata = {
                        elem.keyword: str(elem.value)
                        for elem in ds if elem.VR != "SQ"
                    }

                    relative_path = os.path.relpath(file_path, tmp_dir)
                    gridfs_path = f"{relative_path}"

                    with open(file_path, "rb") as f:
                        file_id = await fs_bucket.upload_from_stream(file_name, f, metadata=metadata)

                    await files_collection.find_one_and_update(
                        {"_id": file_id},
                        {"$set": {"filename": gridfs_path}}
                    )
                    uploaded_files.append(gridfs_path)
                except Exception as e:
                    print(f"❌ Failed {file_path}: {e}")

        return {
            "message": "Patient uploaded successfully",
            "patient_id": patient_id,
            "diagnosis": diagnosis,
            "regimen": regimen,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {e}")